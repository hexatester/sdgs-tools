import logging
from bs4 import BeautifulSoup, Tag
from dateutil.parser import parse as date_parse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Tuple

logger = logging.getLogger(__name__)


def add_header(ws: Worksheet):
    ws["A1"] = "ID"
    ws["B1"] = "Provinsi"
    ws["C1"] = "Kabupaten"
    ws["D1"] = "Kecamatan"
    ws["E1"] = "Desa"
    ws["F1"] = "RT"
    ws["G1"] = "Rw"
    ws["H1"] = "No KK"
    ws["I1"] = "NIK"
    ws["J1"] = "Nama"
    ws["K1"] = "Jenis Kelamin"
    ws["L1"] = "Tgl Lahir"
    ws["M1"] = "Tempat Lahir"
    ws["N1"] = "Usia"
    ws["O1"] = "Status"
    ws["P1"] = "Usia Saat Nikah"
    ws["Q1"] = "Agama"
    ws["R1"] = "Suku Bangsa"
    ws["S1"] = "Warga Negara"
    ws["T1"] = "No HP"
    ws["U1"] = "No WA"
    ws["V1"] = "Email"
    ws["W1"] = "Facebook"
    ws["X1"] = "Instagram"
    ws["Y1"] = "Twitter"
    ws["Z1"] = "Aktif Internet"
    ws["AA1"] = "Akses Internet"
    ws["AB1"] = "Kecepatan Internet"
    ws["AC1"] = "User ID"
    ws["AD1"] = "Tgl Entri"
    ws["AE1"] = "Ter Upload"


def rt_rw(data: str) -> Tuple[str, str]:
    data = data.strip("\xa0")
    if "/" in data:
        data = data.replace("0", "")
        return tuple(data.split("/"))  # type: ignore
    if " " in data:
        data = data.replace("0", "")
        return tuple(data.split(" "))  # type: ignore
    if len(data) == 3 and "0" in data:
        return tuple(data.split("0"))  # type: ignore
    elif len(data) == 4 and data.count("0") == 2:
        res = data.split("0")
        if len(res) == 3:
            return (res[1], res[2])
    return data, ""


def add_row(ws: Worksheet, tds: List[Tag], row: int = 2):
    nik: str = tds[7].get_text()
    nama: str = tds[8].get_text()
    logger.debug(f"Menambahkan data {nik}/{nama}")
    ws[f"A{row}"] = tds[0].get_text()
    ws[f"B{row}"] = tds[1].get_text()
    ws[f"C{row}"] = tds[2].get_text()
    ws[f"D{row}"] = tds[3].get_text()
    ws[f"E{row}"] = tds[4].get_text()
    ws[f"F{row}"], ws[f"G{row}"] = rt_rw(tds[5].get_text())
    ws[f"H{row}"] = tds[6].get_text()
    ws[f"I{row}"] = nik
    ws[f"J{row}"] = nama
    ws[f"K{row}"] = tds[9].get_text()
    ws[f"L{row}"] = date_parse(tds[10].get_text()).date()
    ws[f"M{row}"] = tds[11].get_text()
    ws[f"N{row}"] = tds[12].get_text()
    ws[f"O{row}"] = tds[13].get_text()
    ws[f"P{row}"] = tds[14].get_text()
    ws[f"Q{row}"] = tds[15].get_text()
    ws[f"R{row}"] = tds[16].get_text()
    ws[f"S{row}"] = tds[17].get_text()
    ws[f"T{row}"] = tds[18].get_text()
    ws[f"U{row}"] = tds[19].get_text()
    ws[f"V{row}"] = tds[20].get_text()
    ws[f"W{row}"] = tds[21].get_text()
    ws[f"X{row}"] = tds[22].get_text()
    ws[f"Y{row}"] = tds[23].get_text()
    ws[f"Z{row}"] = tds[24].get_text()
    ws[f"AA{row}"] = tds[25].get_text()
    ws[f"AB{row}"] = tds[26].get_text()
    ws[f"AC{row}"] = tds[27].get_text()
    ws[f"AD{row}"] = tds[28].get_text()
    ws[f"AE{row}"] = tds[29].get_text()


def html_to_xlsx(source: str, destination: str, offset: int = 2):
    try:
        with open(source, "r") as sumber:
            soup = BeautifulSoup(sumber.read(), "html.parser")
    except Exception as e:
        logger.error(f"Gagal membuka file {source} karena {e}")
    try:
        table: Tag = soup.find("table")
        data: List[Tag] = table.find_all("tr")
        data = data[1:]
    except Exception as e:
        logger.error(
            "Format file tidak valid, silahkan download ulang dari api sdgs "
            "atau hubungi pengembang (https://t.me/hexatester)"
        )
    wb = Workbook()
    ws = wb.active
    add_header(ws)
    for index, row in enumerate(data):
        try:
            add_row(ws, row.find_all("td"), index + offset)
        except Exception as e:
            logger.error(f"Gagal menambahkan data baris ke {index+1}, karena {e}")
    try:
        wb.save(destination)
    except Exception as e:
        logger.error(f"Gagal menyimpan file karena {e}")
    logger.info(f"Berhasil mengeksport data sebanyak {len(data)}")
