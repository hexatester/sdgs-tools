from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from uiautomator2 import Device, UiObject

from sdgs_tools.aplikasi_sdgs.utils import d_get_text, menu_to

# PER BOX! resourceId
PENGHASILAN_COL = {
    # "A": "NIK",
    "B": "com.kemendes.survey:id/txtSumber",
    "C": "com.kemendes.survey:id/txtJumlah",
    "D": "com.kemendes.survey:id/txtSatuan",
    "E": "com.kemendes.survey:id/txtPenghasilanSetahun",
    "F": "com.kemendes.survey:id/txtDiekspor",
    "G": "com.kemendes.survey:id/txtStatus",
}


def get_data_penghasilan(d: Device, ws: Worksheet, nik: str, row: int):
    menu_to(d, "PENGHASILAN")
    # d(text="PENGHASILAN").click()
    box_daftar_penghasilan = d(resourceId="com.kemendes.survey:id/itemsPenghasilan")
    survey_box: UiObject = box_daftar_penghasilan.child(
        resourceId="com.kemendes.survey:id/box"
    )
    if survey_box.count == 0:
        return row
    for box_penghasilan in survey_box:
        ws[f"A{row}"] = nik
        # Sumber penghasilan
        box_content = box_penghasilan.child(
            resourceId="com.kemendes.survey:id/txtSumber"
        )
        if box_content.exists:
            ws[f"B{row}"] = box_content.info.get("text")
        # Jumlah
        box_content = box_penghasilan.child(
            resourceId="com.kemendes.survey:id/txtJumlah"
        )
        if box_content.exists:
            ws[f"C{row}"] = box_content.info.get("text")
        # Satuan
        box_content = box_penghasilan.child(
            resourceId="com.kemendes.survey:id/txtSatuan"
        )
        if box_content.exists:
            ws[f"D{row}"] = box_content.info.get("text")
        # Penghasilan Setahun
        box_content = box_penghasilan.child(
            resourceId="com.kemendes.survey:id/txtPenghasilanSetahun"
        )
        if box_content.exists:
            value: str = box_content.info.get("text")
            ws[f"E{row}"] = value.lstrip("Penghasilan setahun : ")
        # Diekspor
        box_content = box_penghasilan.child(
            resourceId="com.kemendes.survey:id/txtDiekspor"
        )
        if box_content.exists:
            value = box_content.info.get("text")
            ws[f"F{row}"] = value.lstrip("Ekspor : ")
        # Status
        box_content = box_penghasilan.child(
            resourceId="com.kemendes.survey:id/txtStatus"
        )
        if box_content.exists:
            ws[f"G{row}"] = box_content.info.get("text")
        row += 1
    d(className='android.widget.ScrollView').fling.vert.backward()
    return row
