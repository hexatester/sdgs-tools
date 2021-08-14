from uiautomator2 import Device, UiObjectNotFoundError
from openpyxl.worksheet.worksheet import Worksheet

from sdgs_tools.aplikasi_sdgs.utils import d_get_text

INDIVIDU_COL = {
    "D": "com.kemendes.survey:id/txtNama",
    "E": "com.kemendes.survey:id/cbJenisKelamin",
    "F": "com.kemendes.survey:id/txtTempatLahir",
    "G": "com.kemendes.survey:id/txtTglLahir",
    "H": "com.kemendes.survey:id/txtUsia",
    "I": "com.kemendes.survey:id/cbStatus",
    "J": "com.kemendes.survey:id/txtUsiaSaatNikah",
    "K": "com.kemendes.survey:id/cbAgama",
    "L": "com.kemendes.survey:id/txtSuku",
    "M": "com.kemendes.survey:id/cbWargaNegara",
    "N": "com.kemendes.survey:id/txtNoHP",
    "O": "com.kemendes.survey:id/txtNoWA",
    "P": "com.kemendes.survey:id/txtEmail",
    "Q": "com.kemendes.survey:id/txtFB",
    "R": "com.kemendes.survey:id/txtTwitter",
    "S": "com.kemendes.survey:id/txtInstagram",
    "T": "com.kemendes.survey:id/cbInternet",
    # "U": "com.kemendes.survey:id/cbAksesInternetLewat",
    # "V": "com.kemendes.survey:id/cbKecepatanInternet",
}


def get_data_individu(d: Device, ws: Worksheet, row: int):
    d(text="DATA INDIVIDU").click()
    for col, resourceId in INDIVIDU_COL.items():
        cell = f"{col}{row}"
        try:
            ws[cell] = d_get_text(d, resourceId)
        except UiObjectNotFoundError:
            continue
    if ws[f"T{row}"].value == "Ya":
        # Internet = Ya
        ws[f"U{row}"] = d_get_text(d, "com.kemendes.survey:id/cbAksesInternetLewat")
        ws[f"V{row}"] = d_get_text(d, "com.kemendes.survey:id/cbKecepatanInternet")
