import cattr
import click
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from time import sleep
from typing import Any, Dict, List, Tuple
from uiautomator2 import Device

from sdgs_tools.aplikasi_sdgs.utils import menu_to
from sdgs_tools.dashboard.import_individu import DataIndividu
from sdgs_tools.dashboard.import_individu.mapping import MAPPING_COLS

from .individu import get_data_individu
from .kesehatan import get_data_kesehatan
from .pekerjaan import get_data_pekerjaan
from .pendidikan import get_data_pendidikan
from .penghasilan import get_data_penghasilan


__all__ = [
    "get_data_individu",
    "get_data_kesehatan",
    "get_data_pekerjaan",
    "get_data_pendidikan",
    "get_data_penghasilan",
]


def get_param(
    ws: Worksheet, row: int, kk: str = "rt_rw", nik: str = "nik"
) -> Tuple[str, str]:
    kk_col = MAPPING_COLS[kk]
    nik_col = MAPPING_COLS[nik]
    return (ws[f"{kk_col}{row}"].value, ws[f"{nik_col}{row}"].value)


def export_individu(
    d: Device,
    filepath: str,
    rt_rw: str,
    rows: List[int],
    start_row_penghasilan: int = 2,
    skip_individu: bool = False,
    skip_pekerjaan: bool = False,
    skip_penghasilan: bool = False,
    skip_kesehatan: bool = False,
    skip_disabilitas: bool = False,
    skip_pendidikan: bool = False,
):
    wb = load_workbook(filepath)
    individu_ws = wb["Individu"]
    form_rt_rw = d(resourceId="com.kemendes.survey:id/txtRTRW")
    while not form_rt_rw.exists:
        click.echo(
            "Mohon aktifkan GPS, buka aplikasi, dan masuk menu Entri Survey Individu!"
        )
        sleep(1)
    skipped = 0
    failed = 0
    success = 0
    row_penghasilan = start_row_penghasilan
    for row in rows:
        no_kk, nik = get_param(individu_ws, row)
        # Input rt rw no_kk nik & Tampilkan
        form_rt_rw = d(resourceId="com.kemendes.survey:id/txtRTRW")
        form_rt_rw.send_keys(rt_rw)
        form_kk = d(resourceId="com.kemendes.survey:id/txtNoKK")
        form_kk.send_keys(no_kk)
        form_nik = d(resourceId="com.kemendes.survey:id/txtNIK")
        form_nik.send_keys(nik)
        d.press("back")
        d(resourceId="com.kemendes.survey:id/btnCariRT").click()
        menu_to(d, "DATA INDIVIDU")
        # Get Data
        data: Dict[str, Any] = dict()
        # Individu
        if not skip_individu:
            data.update(get_data_individu(d))
        # Pekerjaan
        if not skip_pekerjaan:
            data.update(get_data_pekerjaan(d))
        # Penghasilan
        if not skip_penghasilan:
            data["penghasilan"] = get_data_penghasilan(d)
        # Kesehatan
        if not skip_kesehatan:
            data.update(get_data_kesehatan(d))
        # Pendidikan
        if not skip_pendidikan:
            data.update(get_data_pendidikan(d))
        try:
            individu: DataIndividu = cattr.structure(data, DataIndividu)
            row_penghasilan = individu.save(wb, row, row_penghasilan)
            success += 1
        except ValueError as e:
            click.echo(f"Baris {row} dilewati karena : {e}")
            skipped += 1
            continue
        except Exception as e:
            click.echo(f"Error ketika membuat DataKeluarga baris {row} : {e}")
            failed += 1
            continue
