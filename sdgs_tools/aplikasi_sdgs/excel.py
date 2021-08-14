from openpyxl import Workbook

from .utils import set_ws_header

INDIVIDU = {
    "A": "RT/RW",
    "B": "Nomor KK",
    "C": "NIK",
    "D": "Nama",
    "E": "Jenis Kelamin",
    "F": "Tempat Lahir",
    "G": "Tgl Lahir",
    "H": "Usia",
    "I": "Status",
    "J": "Usia Menikah",
    "K": "Agama",
    "L": "Suku",
    "M": "Warga Negara",
    "N": "No HP",
    "O": "No WA",
    "P": "Email",
    "Q": "FB",
    "R": "Twitter",
    "S": "Instagram",
    "T": "Internet",
    "U": "Akses Internet Lewat",
    "V": "Kecepatan Internet",
    # Pekerjaan
    "W": "Kondisi Pekerjaan",
    "X": "Pekerjaan Utama",
    "Y": "Jaminan Sosial",
    "Z": "Penghasilan",
    # Kesehatan - Penyakit
    "AA": "Muntaber",
    "AB": "Demam Berdarah",
    "AC": "Campak",
    "AC": "Malaria",
    "AE": "Flu Burung",
    "AF": "Covid-19",
    "AG": "Hepatitis B",
    "AH": "Leptospirosis",
    "AI": "Kolera",
    "AJ": "Gizi Buruk",
    "AK": "Jantung",
    "AL": "TBC Paru",
    "AM": "Kanker",
    "AN": "Diabetes",
    "AO": "Hepatitis E",
    "AP": "Difteri",
    "AQ": "Chikungunya",
    "AR": "Lumpuh",
    "AS": "Lainnya",
    # Kesehatan - Fasilitas kesehatan
    "AT": "RS",
    "AU": "RS Bersalin",
    "AV": "Puskesmas Rawat Inap",
    "AW": "Puskesmas Tanpa Rawat Inap",
    "AX": "Puskesmas Pembantu",
    "AY": "Poliklinik",
    "AZ": "Praktik Dokter",
    "BA": "Rumah Bersalin",
    "BB": "Praktik Bidan",
    "BC": "Poskedes",
    "BC": "Polindes",
    "BE": "Apotik",
    "BF": "Toko Khusus",
    "BG": "Posyandu",
    "BH": "Posbindu",
    "BI": "Praktik Dukun",
    "BJ": "Jaminan Sosial Kesehatan",
    "BK": "Melahirkan",
    # Disabilitas
    "BL": "Buta",
    "BM": "Tuli",
    "BN": "Bisu",
    "BO": "Bisu-tuli",
    "BP": "Cacat tubuh",
    "BQ": "Cacat mental",
    "BR": "Eks-sakit jiwa",
    "BS": "Eks-sakit kusta",
    "BT": "Cacat ganda",
    "BU": "Dipasung",
    # Pendidikan
    "BV": "Pendidikan Tertinggi",
    "BW": "Tahun Pendidikan Dasar",
    "BX": "Bahasa",
    "BY": "Bahasa Lembaga",
    "BZ": "Kerja bakti",
    "CA": "Siskamling",
    "CB": "Pesta rakyat",
    "CC": "Menolong kematian",
    "CD": "Menolong sakit",
    "CE": "Menolong kecelakaan",
    "CF": "Pelayanan desa",
    "CG": "Masukan",
    "CH": "Bencana",
}

INDIVIDU_PENGHASILAN = {
    "A": "NIK",
    "B": "Sumber",
    "C": "Jumlah",
    "D": "Satuan",
    "E": "PenghasilanSetahun",
    "F": "Diekspor",
    "F": "Status",
}

KELUARGA = {
    "A": "RT/RW",
    "B": "Nomor KK",
    "C": "Nama",
    "D": "Alamat",
    "E": "No. HP",
    "F": "No. Telepon",
    "G": "NIK Kepala Keluarga",
    "H": "Status Tempat Tinggal",
    "I": "Status Lahan",
    "J": "Luas Lantai",
    "K": "Luas Tempat Tinggal",
    "L": "Jenis Lantai",
    "M": "Dinding",
    "N": "Jendela",
    "O": "Atap",
    "P": "Penerangan",
    "Q": "Memasak",
    "R": "Sampah",
    "S": "MCK",
    "T": "Air Mandi",
    "U": "Fasilitas BAB",
    "V": "Air Minum",
    "W": "Pembuangan Air Limbah",
    "X": "Dibawah SUTET",
    "Y": "Di bantaran sungai",
    "Z": "Di lereng gunung / bukit",
    "AA": "Kondisi Rumah",
    # Lain lain
    "AB": "Jumlah Anggota Keluarga Menggunakan Transportasi Umum",
    "AC": "Jumlah Anggota Keluarga Menggunakan Transportasi Umum bulan sebelum",
    "AD": "BLT",
    "AE": "PKH",
    "AF": "BST",
    "AG": "Banpres",
    "AH": "UMKM",
    "AI": "Bantuan Pekerja",
    "AJ": "Pendidikan Anak",
    "AK": "Ban Lainnya",
    "AL": "Rata-rata Pengeluaran",
}

KELUARGA_PENDIDIKAN = {
    "A": "Nomor KK",
    "B": "Fasilitas",
    "C": "Jarak Tempuh",
    "D": "Waktu Tempuh",
    "E": "Kemudahan",
    "F": "Keterangan",
}

KELUARGA_KESEHATAN = {
    "A": "Nomor KK",
    "B": "Fasilitas",
    "C": "Jarak Tempuh",
    "D": "Waktu Tempuh",
    "E": "Kemudahan",
    "F": "Keterangan",
}

KELUARGA_TENAGA_KESEHATAN = {
    "A": "Nomor KK",
    "B": "Fasilitas",
    "C": "Jarak Tempuh",
    "D": "Waktu Tempuh",
    "E": "Kemudahan",
    "F": "Keterangan",
}

KELUARGA_SARPRAS = {
    "A": "Nomor KK",
    "B": "Tujuan",
    "C": "Jenis Transportasi",
    "D": "Transportasi Umum",
    "E": "Waktu Tempuh",
    "F": "Biaya",
    "G": "Kemudahan",
    "H": "Keterangan",
}


def add_header_individu(wb: Workbook, row: int = 1):
    set_ws_header(wb.create_sheet("Individu"), INDIVIDU, row)
    set_ws_header(wb.create_sheet("Penghasilan"), INDIVIDU_PENGHASILAN, row)


def add_header_keluarga(wb: Workbook, row: int):
    set_ws_header(wb.create_sheet("Keluarga"), KELUARGA, row)
    set_ws_header(wb.create_sheet("Pendidikan"), KELUARGA_PENDIDIKAN, row)
    set_ws_header(wb.create_sheet("Kesehatan"), KELUARGA_KESEHATAN, row)
    set_ws_header(wb.create_sheet("Tenaga Kesehatan"), KELUARGA_TENAGA_KESEHATAN, row)
    set_ws_header(wb.create_sheet("Sarpras"), KELUARGA_SARPRAS, row)


def make_template_individu(filepath: str = "Data INDIVIDU SDGS.xlsx", row: int = 1):
    if not filepath.endswith(".xlsx"):
        filepath += ".xlsx"
    wb = Workbook()
    add_header_individu(wb, row)
    wb.save(filepath)


def make_template_keluarga(filepath: str = "Data KELUARGA SDGS.xlsx", row: int = 1):
    if not filepath.endswith(".xlsx"):
        filepath += ".xlsx"
    wb = Workbook()
    add_header_keluarga(wb, row)
    wb.save(filepath)
