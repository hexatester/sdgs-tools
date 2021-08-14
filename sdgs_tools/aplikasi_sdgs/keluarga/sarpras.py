from uiautomator2 import Device, UiObject
from openpyxl.worksheet.worksheet import Worksheet

from sdgs_tools.aplikasi_sdgs.utils import d_get_text, menu_to, swipe_box


SARPRAS_COL = {
    "B": ("com.kemendes.survey:id/txtNIK", ""),
    # "C": ("com.kemendes.survey:id/txtNama", "Tujuan : Lokasi pekerjaan utama Jenis Transportasi : Darat"),
    # "D": ("com.kemendes.survey:id/txtAlamat", "Waktu Tempuh :0.25 jam, Biaya : 5000"),
    "G": ("com.kemendes.survey:id/txtTelpon", "Kemudahan :"),
    "H": ("com.kemendes.survey:id/txtStatus", ""),
}


def get_data_sarpras(d: Device, ws: Worksheet, no_kk: str, row: int):
    menu_to(d, "AKSES SARANA PRASARANA")
    box_daftar_sarpras: UiObject = d(resourceId="com.kemendes.survey:id/itemsSarana")
    survey_boxes: UiObject = box_daftar_sarpras.child(
        resourceId="com.kemendes.survey:id/box"
    )
    if survey_boxes.count == 0:
        return row
    for survey_box in survey_boxes:
        ws[f"A{row}"] = no_kk
        for col, (resourceId, lstrp) in SARPRAS_COL.items():
            current = survey_box.child(resourceId=resourceId)
            value: str = current.info.get("text")
            if isinstance(value, str):
                ws[f"{col}{row}"] = value.lstrip(lstrp)
        tujuan_jenis: str = survey_box.child(resourceId="com.kemendes.survey:id/txtNama").info.get("text")
        tujuan_jenis = tujuan_jenis.lstrip("Tujuan : ")
        tujuan_jenis = tujuan_jenis.replace(" Jenis Transportasi", "")
        tujuan, jenis = tujuan_jenis.split(":")
        ws[f"C{row}"] = tujuan.strip()
        ws[f"D{row}"] = jenis.strip()
        waktu_biaya: str = survey_box.child(resourceId="com.kemendes.survey:id/txtAlamat").info.get("text")
        waktu_biaya = waktu_biaya.lstrip("Waktu Tempuh :")
        waktu_biaya = waktu_biaya.replace(", Biaya", "")
        waktu, biaya = waktu_biaya.split(":")
        ws[f"E{row}"] = waktu.strip()
        ws[f"F{row}"] = biaya.strip()
        row += 1
        swipe_box(d, survey_box)
    d(className="android.widget.ScrollView").fling.vert.backward()
    return row
