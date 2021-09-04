import attr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict

MAPPING = {
    "desa": "desa",
    "I.P103": "nama",
    "I.P104": "jenis_kelamin",
    "I.P105": "tempat_lahir",
    "I.P106": "tanggal_lahir",
    "I.P107": "usia",
    "I.P108": "status_pernikahan",
    "I.P109": "usia_menikah",
    "I.P110": "agama",
    "I.P111": "suku_bangsa",
    "I.P112": "warga_negara",
    "I.P113": "nomor_hp",
    "I.P114": "nomor_whatsapp",
    "I.P115": "alamat_email",
    "I.P116": "alamat_facebook",
    "I.P117": "alamat_twitter",
    "I.P118": "alamat_instagram",
    "I.P119": "aktif_internet",
    "I.P120": "akses_melalui",
    "I.P121": "kecepatan_internet",
    "I.P201": "kondisi_pekerjaan",
    "I.P202": "pekerjaan_utama",
    "I.P202-Comment": "pekerjaan_utama_comment",
    "I.P203": "jaminan_sosial_ketenagakerjaan",
    "I.P204": "penghasilan",
    "I.P204_penghasilan": "pekerjaan_penghasilan",
    "I.P401": "penyakit_diderita",
    "I.P402": "fasilitas_kesehatan",
    "I.P403": "jamsos_ketenagakerjaan",
    "I.P404": "penyakit_diderita",
    "I.P405": "setahun_melahirkan",
    "I.P406": "mendapat_asi",
    "I.P501": "pendidikan_tertinggi",
    "I.P502": "tahun_pendidikan",
    "I.P505": "bahasa_permukiman",
    "I.P506": "bahasa_formal",
    "I.P507": "kerja_bakti",
    "I.P508": "siskamling",
    "I.P509": "pesta_rakyat",
    "I.P510": "menolong_kematian",
    "I.P511": "menolong_sakit",
    "I.P512": "menolong_kecelakaan",
    "I.P513": "memperoleh_pelayanan_desa",
    "I.P514": "pelayanan_desa",
    "I.P515": "saran_desa",
    "I.P516": "keterbukaan_desa",
    "I.P517": "terjadi_bencana",
    "I.P518": "terdampak_bencana",
    "kecamatan": "kecamatan",
    "kota": "kota",
    "nik": "nik",
    "no_kk": "no_kk",
    "provinsi": "provinsi",
    "rt": "rt",
    "rw": "rw",
}


@attr.dataclass
class MappingIndividu:
    no_kk: str = "A"
    nik: str = "B"
    nama: str = "C"
    jenis_kelamin: str = "D"
    tempat_lahir: str = "E"
    tanggal_lahir: str = "F"
    usia: str = "G"
    status_pernikahan: str = "H"
    agama: str = "I"
    suku: str = "J"
    warganegara: str = "K"
    nomor_hp: str = "L"
    aktif_internet: str = "M"
    kecepatan_internet: str = "N"
    agama_comment: str = "O"
    akses_melalui: str = "P"
    usia_menikah: str = "Q"
    nomor_whatsapp: str = "R"
    alamat_email: str = "S"
    alamat_facebook: str = "T"
    alamat_twitter: str = "U"
    alamat_instagram: str = "V"

    def make_row(self, ws: Worksheet, row: int) -> Dict[str, Any]:
        return {
            "no_kk": ws[f"{self.no_kk}{row}"],
            "nik": ws[f"{self.nik}{row}"],
            "nama": ws[f"{self.nama}{row}"],
            "jenis_kelamin": ws[f"{self.jenis_kelamin}{row}"],
            "tempat_lahir": ws[f"{self.tempat_lahir}{row}"],
            "tanggal_lahir": ws[f"{self.tanggal_lahir}{row}"],
            "usia": ws[f"{self.usia}{row}"],
            "status_pernikahan": ws[f"{self.status_pernikahan}{row}"],
            "agama": ws[f"{self.agama}{row}"],
            "suku": ws[f"{self.suku}{row}"],
            "warganegara": ws[f"{self.warganegara}{row}"],
            "nomor_hp": ws[f"{self.nomor_hp}{row}"],
            "aktif_internet": ws[f"{self.aktif_internet}{row}"],
            "kecepatan_internet": ws[f"{self.kecepatan_internet}{row}"],
            "agama_comment": ws[f"{self.agama_comment}{row}"],
            "akses_melalui": ws[f"{self.akses_melalui}{row}"],
            "usia_menikah": ws[f"{self.usia_menikah}{row}"],
            "nomor_whatsapp": ws[f"{self.nomor_whatsapp}{row}"],
            "alamat_email": ws[f"{self.alamat_email}{row}"],
            "alamat_facebook": ws[f"{self.alamat_facebook}{row}"],
            "alamat_twitter": ws[f"{self.alamat_twitter}{row}"],
            "alamat_instagram": ws[f"{self.alamat_instagram}{row}"],
        }
