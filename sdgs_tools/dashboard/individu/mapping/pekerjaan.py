import attr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict


@attr.dataclass
class MappingPekerjaan:
    kondisi_pekerjaan: str = "X"
    pekerjaan_utama: str = "Y"
    pekerjaan_utama_comment: str = "Z"
    jaminan_sosial_ketenagakerjaan: str = "AA"

    def make_row(self, ws: Worksheet, row: int) -> Dict[str, Any]:
        return {
            "kondisi_pekerjaan": ws[f"{self.kondisi_pekerjaan}{row}"].value,
            "pekerjaan_utama": ws[f"{self.pekerjaan_utama}{row}"].value,
            "pekerjaan_utama_comment": ws[f"{self.pekerjaan_utama_comment}{row}"].value,
            "jaminan_sosial_ketenagakerjaan": ws[
                f"{self.jaminan_sosial_ketenagakerjaan}{row}"
            ].value,
        }
