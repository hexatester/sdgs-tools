import attr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict


@attr.dataclass
class MappingIndividu:
    no_kk: str = "A"
    nik: str = "B"
    nama: str = "C"
    jenis_kelamin: str = "D"
    tempat_lahir: str = "E"
    tanggal_lahir: str = "F"
    usia: str = "G"
    status_pernikahan: str = "H"
    agama: str = "I"
    suku: str = "J"
    warganegara: str = "K"
    nomor_hp: str = "L"
    aktif_internet: str = "M"
    kecepatan_internet: str = "N"
    agama_comment: str = "O"
    akses_melalui: str = "P"
    usia_menikah: str = "Q"
    nomor_whatsapp: str = "R"
    alamat_email: str = "S"
    alamat_facebook: str = "T"
    alamat_twitter: str = "U"
    alamat_instagram: str = "V"

    def make_row(self, ws: Worksheet, row: int) -> Dict[str, Any]:
        return {
            "no_kk": ws[f"{self.no_kk}{row}"].value,
            "nik": ws[f"{self.nik}{row}"].value,
            "nama": ws[f"{self.nama}{row}"].value,
            "jenis_kelamin": ws[f"{self.jenis_kelamin}{row}"].value,
            "tempat_lahir": ws[f"{self.tempat_lahir}{row}"].value,
            "tanggal_lahir": ws[f"{self.tanggal_lahir}{row}"].value,
            "usia": ws[f"{self.usia}{row}"].value,
            "status_pernikahan": ws[f"{self.status_pernikahan}{row}"].value,
            "agama": ws[f"{self.agama}{row}"].value,
            "suku": ws[f"{self.suku}{row}"].value,
            "warganegara": ws[f"{self.warganegara}{row}"].value,
            "nomor_hp": ws[f"{self.nomor_hp}{row}"].value,
            "aktif_internet": ws[f"{self.aktif_internet}{row}"].value,
            "kecepatan_internet": ws[f"{self.kecepatan_internet}{row}"].value,
            "agama_comment": ws[f"{self.agama_comment}{row}"].value,
            "akses_melalui": ws[f"{self.akses_melalui}{row}"].value,
            "usia_menikah": ws[f"{self.usia_menikah}{row}"].value,
            "nomor_whatsapp": ws[f"{self.nomor_whatsapp}{row}"].value,
            "alamat_email": ws[f"{self.alamat_email}{row}"].value,
            "alamat_facebook": ws[f"{self.alamat_facebook}{row}"].value,
            "alamat_twitter": ws[f"{self.alamat_twitter}{row}"].value,
            "alamat_instagram": ws[f"{self.alamat_instagram}{row}"].value,
        }
