import attr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict, List


@attr.dataclass
class MappingPenghasilan:
    nik: str = "A"
    diekspor: str = "B"
    jumlah: str = "C"
    penghasilan: str = "D"
    sumber_penghasilan: str = "E"
    comment: str = "F"

    def make_penghasilan(
        self,
        ws: Worksheet,
        nik: str,
        last_row: int = 2,
        max_offset: int = 25,
    ) -> List[Dict[str, Any]]:
        next_row_offset: int = 0
        results: List[Dict[str, Any]] = list()
        while next_row_offset < max_offset:
            row = last_row + next_row_offset
            row_nik = ws[f"{self.nik}{row}"].value
            if row_nik == nik:
                results.append(
                    {
                        "diekspor": ws[f"{self.diekspor}{row}"].value,
                        "jumlah": ws[f"{self.jumlah}{row}"].value,
                        "penghasilan": ws[f"{self.penghasilan}{row}"].value,
                        "sumber_penghasilan": ws[
                            f"{self.sumber_penghasilan}{row}"
                        ].value,
                        "comment": ws[f"{self.comment}{row}"].value,
                    }
                )
        return results
