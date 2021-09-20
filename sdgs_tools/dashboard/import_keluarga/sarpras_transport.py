import attr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional

from .enums import JenisTransportasi, YaTidak


@attr.dataclass
class SarprasTransport:
    jenis: Optional[JenisTransportasi] = None
    transport_umum: Optional[YaTidak] = None
    waktu: Optional[str] = None
    biaya: Optional[str] = None
    kemudahan: Optional[YaTidak] = None

    def __attrs_post_init__(self):
        if self.waktu in (None, "None"):
            self.waktu = "0"
        elif self.waktu.isdigit():
            waktu = int(self.waktu) / 60
            self.waktu = str(round(waktu, 4))
        elif self.waktu.isdecimal():
            waktu = float(self.waktu) / 60
            self.waktu = str(round(waktu, 4))
        if not self.biaya:
            self.biaya = 0

    @staticmethod
    def from_cols(
        ws: Worksheet,
        row: int,
        j: str,
        t: str,
        w: str,
        b: str,
        k: str,
    ):
        return {
            "jenis": ws[f"{j}{row}"].value,
            "transport_umum": ws[f"{t}{row}"].value,
            "waktu": ws[f"{w}{row}"].value,
            "biaya": ws[f"{b}{row}"].value,
            "kemudahan": ws[f"{k}{row}"].value,
        }

    def save(
        self,
        ws: Worksheet,
        row: int,
        j: str,
        t: str,
        w: str,
        b: str,
        k: str,
    ):
        if self.jenis is not None:
            ws[f"{j}{row}"] = self.jenis.value
        if self.transport_umum is not None:
            ws[f"{t}{row}"] = self.transport_umum.value
        if self.waktu is not None:
            ws[f"{w}{row}"] = self.waktu
        if self.biaya is not None:
            ws[f"{b}{row}"] = self.biaya
        if self.kemudahan is not None:
            ws[f"{k}{row}"] = self.kemudahan.value
