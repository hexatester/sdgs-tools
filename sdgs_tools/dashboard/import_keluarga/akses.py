import attr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict, Optional

from .enums import YaTidak


@attr.dataclass
class Akses:
    jarak: Optional[float] = None
    waktu: Optional[float] = None
    kemudahan: Optional[YaTidak] = None

    @staticmethod
    def from_cols(ws: Worksheet, row: int, j: str, w: str, k: str):
        return {
            "jarak": ws[f"{j}{row}"].value,
            "waktu": ws[f"{w}{row}"].value,
            "kemudahan": ws[f"{k}{row}"].value,
        }

    def todict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {
            "jarak": None,
            "waktu": None,
            "kemudahan": None,
        }
        if isinstance(self.jarak, float):
            if self.waktu == 0:
                data["jarak"] = "0"
            else:
                data["jarak"] = str(self.jarak)
        elif isinstance(self.jarak, int):
            data["jarak"] = str(self.jarak)
        if isinstance(self.waktu, float):
            if self.waktu == 0:
                data["waktu"] = "0"
            else:
                waktu = self.waktu / 60
                data["waktu"] = str(round(waktu, 4))
        elif isinstance(self.waktu, int):
            waktu = self.waktu / 60
            data["waktu"] = str(round(waktu, 4))
        if self.kemudahan is not None:
            data["kemudahan"] = self.kemudahan.value
        return data

    def save(self, ws: Worksheet, row: int, j: str, w: str, k: str):
        if self.jarak is not None:
            ws[f"{j}{row}"] = self.jarak
        if self.waktu is not None:
            ws[f"{w}{row}"] = self.waktu
        if self.kemudahan is not None:
            ws[f"{k}{row}"] = self.kemudahan.value
