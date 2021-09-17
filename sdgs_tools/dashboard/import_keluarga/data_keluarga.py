import attr
import cattr
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict, Optional, Tuple

from . import (
    Luas,
    AksesFasilitasKesehatan,
    AksesPendidikan,
    AksesSarprasTransport,
    AksesTenagaKesehatan,
    BantuanPemerintah,
)
from .enums import (
    Atap,
    Dinding,
    EnergiMemasak,
    FasilitasBab,
    FasilitasMck,
    Jendela,
    Lantai,
    PembuanganLimbahCair,
    PembuanganSampah,
    Penerangan,
    StatusLahan,
    SumberAirMandi,
    SumberAirMinum,
    SumberKayuBakar,
    TempatTinggal,
    YaTidak,
)
from .mapping import MAPPING, MAPPING_COLS


@attr.dataclass(kw_only=True)
class DataKeluarga:
    # DESKRIPSI KELUARGA & PEMUKIMAN
    no_kk: str
    nik: str
    nama: str
    alamat: str
    nomor_hp: str
    telepon_rumah: str
    anggota_keluarga: str
    tempat_tinggal: TempatTinggal
    tempat_tinggal_comment: Optional[str] = None
    status_lahan: StatusLahan
    status_lahan_comment: Optional[str] = None
    luas: Luas
    lantai: Lantai
    lantai_other: Optional[str] = None
    dinding: Dinding
    dinding_other: Optional[str] = None
    jendela: Jendela
    atap: Atap
    atap_comment: Optional[str] = None
    penerangan: Penerangan
    energi_memasak: EnergiMemasak
    sumber_kayu_bakar: Optional[SumberKayuBakar] = None
    pembuangan_sampah: PembuanganSampah
    fasilitas_mck: FasilitasMck
    sumber_air_mandi: SumberAirMandi
    sumber_air_mandi_comment: Optional[str] = None
    fasilitas_bab: FasilitasBab
    fasilitas_bab_comment: Optional[str] = None
    sumber_air_minum: SumberAirMinum
    sumber_air_minum_comment: Optional[str] = None
    pembuangan_limbah_cair: PembuanganLimbahCair
    pembuangan_limbah_cair_comment: Optional[str] = None
    bawah_sutet: YaTidak
    bantaran_sungai: YaTidak
    lereng_gunung: YaTidak
    kumuh: YaTidak
    # AKSES PENDIDIKAN TERDEKAT
    akses_pendidikan: AksesPendidikan
    # AKSES FASILITAS KESEHATAN TERDEKAT
    akses_fasilitas_kesehatan: AksesFasilitasKesehatan
    # AKSES TENAGA KESEHATAN TERDEKAT
    akses_tenaga_kesehatan: AksesTenagaKesehatan
    # AKSES PRASARANA DAN SARANA TRANSPORTASI
    akses_sarpras_transport: AksesSarprasTransport
    # LAIN-LAIN
    transport_umum: str = "0"
    transport_umum_bulan_sebelumnya: str = "0"
    penerima_program_pemerintah: BantuanPemerintah
    pengeluaran_bulanan: str = "0"

    def __attrs_post_init__(self) -> None:
        if (
            self.tempat_tinggal == TempatTinggal.LAINNYA
            and not self.tempat_tinggal_comment
        ):
            raise ValueError("Tempat Tinggal Lainnya harus diisi")
        if self.status_lahan == StatusLahan.LAINNYA and not self.status_lahan_comment:
            raise ValueError("Status Lahan Lainnya harus diisi")
        if self.lantai == Lantai.LAINNYA and not self.lantai_other:
            raise ValueError("Lantai Lainnya harus diisi")
        if self.dinding == Dinding.LAINNYA and not self.dinding_other:
            raise ValueError("Dinding Lainnya harus diisi")
        if self.atap == Atap.LAINNYA and not self.atap_comment:
            raise ValueError("Atap Lainnya harus diisi")
        if (
            self.sumber_air_mandi == SumberAirMandi.LAINNYA
            and not self.sumber_air_mandi_comment
        ):
            raise ValueError("Sumber Air Mandi Lainnya harus diisi")
        if (
            self.fasilitas_bab == FasilitasBab.LAINNYA
            and not self.fasilitas_bab_comment
        ):
            raise ValueError("Fasilitas BAB Lainnya harus diisi")
        if (
            self.sumber_air_minum == SumberAirMinum.LAINNYA
            and not self.sumber_air_minum_comment
        ):
            raise ValueError("Sumber Air Minum Lainnya harus diisi")
        if (
            self.pembuangan_limbah_cair == PembuanganLimbahCair.LAINNYA
            and not self.pembuangan_limbah_cair_comment
        ):
            raise ValueError("Pembuangan Limbah Cair Lainnya harus diisi")

    @staticmethod
    def make(ws: Worksheet, row: int):
        data: Dict[str, Any] = dict()
        for name, col in MAPPING_COLS.items():
            data[name] = ws[f"{col}{row}"].value
        data["luas"] = Luas.make(ws, row)
        data["akses_pendidikan"] = AksesPendidikan.make(ws, row)
        data["akses_fasilitas_kesehatan"] = AksesFasilitasKesehatan.make(ws, row)
        data["akses_tenaga_kesehatan"] = AksesTenagaKesehatan.make(ws, row)
        data["akses_sarpras_transport"] = AksesSarprasTransport.make(ws, row)
        return data

    @staticmethod
    def get_kk_nik(ws: Worksheet, row: int) -> Tuple[str, str]:
        return (
            ws[f"{MAPPING_COLS['no_kk']}{row}"].value,
            ws[f"{MAPPING_COLS['nik']}{row}"].value,
        )

    def make_data(self, desa: str, rt: str, rw: str):
        raw_data: Dict[str, Any] = cattr.unstructure(self)
        clean_data: Dict[str, Any] = dict()
        for key, properti in MAPPING.items():
            value = raw_data.get(properti)
            if value in (None, "None"):
                # TODO Improve empty value detection
                continue
            clean_data[key] = value
        clean_data.update(self.akses_pendidikan.todict())
        clean_data.update(self.akses_tenaga_kesehatan.todict())
        clean_data.update(self.akses_fasilitas_kesehatan.todict())
        clean_data.update(self.akses_sarpras_transport.todict())
        clean_data.update(
            {
                "desa": desa,
                "kecamatan": desa[0:7],
                "kota": desa[0:4],
                "provinsi": desa[0:2],
                "rt": rt,
                "rw": rw,
            }
        )
        return clean_data
